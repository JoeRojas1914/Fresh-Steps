from flask import Blueprint, render_template, request, redirect, flash, jsonify, session


from services.servicios_service import (
    listar_servicios,
    guardar_servicio_service,
    eliminar_servicio_service,
    obtener_historial_servicio_service,
    restaurar_servicio_service
)
from negocio import obtener_negocios

servicios_bp = Blueprint("servicios", __name__)

@servicios_bp.route("/servicios")
def servicios():
    q = request.args.get("q", "")
    id_negocio = request.args.get("id_negocio", type=int)
    pagina = request.args.get("pagina", 1, type=int)

    incluir_eliminados = bool(request.args.get("eliminados"))

    data = listar_servicios(
        id_negocio=id_negocio,
        q=q,
        pagina=pagina,
        por_pagina=10,
        incluir_eliminados=incluir_eliminados 
    )

    negocios = obtener_negocios()

    return render_template(
        "servicios.html",
        servicios=data["servicios"],
        negocios=negocios,
        id_negocio=id_negocio,
        q=q,
        pagina=pagina,
        total_paginas=data["total_paginas"],
        incluir_eliminados=incluir_eliminados 
    )



@servicios_bp.route("/servicios/guardar", methods=["POST"])
def guardar_servicio():
    id_servicio = request.form.get("id_servicio")
    id_negocio = request.form["id_negocio"]
    nombre = request.form["nombre"]
    precio = request.form["precio"]
    id_usuario = session.get("id_usuario")

    resultado = guardar_servicio_service(
        id_servicio,
        id_negocio,
        nombre,
        precio,
        id_usuario
    )


    if resultado == "actualizado":
        flash("✅ Servicio actualizado correctamente.", "success")
    else:
        flash("✅ Servicio creado correctamente.", "success")

    return redirect("/servicios")


@servicios_bp.route("/servicios/eliminar/<int:id_servicio>")
def eliminar_servicio(id_servicio):

    id_usuario = session.get("id_usuario")
    ok = eliminar_servicio_service(id_servicio, id_usuario)


    if not ok:
        flash("❌ No puedes eliminar el servicio porque ya tiene ventas.", "error")
    else:
        flash("✅ Servicio eliminado correctamente.", "success")

    return redirect("/servicios")




@servicios_bp.route("/api/servicios")
def api_servicios():
    id_negocio = request.args.get("id_negocio", type=int)

    data = listar_servicios(
        id_negocio=id_negocio,
        pagina=1,
        por_pagina=1000
    )

    return jsonify(data["servicios"])



@servicios_bp.route("/servicios/<int:id_servicio>/historial")
def historial_servicio(id_servicio):
    data = obtener_historial_servicio_service(id_servicio)
    return jsonify(data)


@servicios_bp.route("/servicios/restaurar/<int:id_servicio>")
def restaurar_servicio(id_servicio):
    id_usuario = session.get("id_usuario")

    restaurar_servicio_service(id_servicio, id_usuario)

    flash("♻️ Servicio restaurado correctamente.", "success")
    return redirect(request.referrer or "/servicios")



@servicios_bp.route("/servicios/exportar")
def exportar_servicios_excel():
    import io
    from datetime import datetime
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from servicios import obtener_servicios

    id_negocio        = request.args.get("id_negocio") or None
    incluir_eliminados = request.args.get("eliminados") == "1"

    servicios = obtener_servicios(
        id_negocio=id_negocio,
        incluir_eliminados=incluir_eliminados,
        limit=99999, offset=0
    )

    C = {
        "azul":     "1E7FD6", "azul_cl":  "E6F3FF",
        "blanco":   "FFFFFF", "gris":     "F8FBFF",
        "verde":    "22C55E", "verde_cl": "DCFCE7",
        "rojo":     "E53935", "rojo_cl":  "FEE2E2",
        "gris_txt": "6B7280", "dark":     "1F2937",
    }
    borde = Border(
        left=Side(style="thin", color="CFD8E3"),
        right=Side(style="thin", color="CFD8E3"),
        top=Side(style="thin", color="CFD8E3"),
        bottom=Side(style="thin", color="CFD8E3"),
    )

    def cell(ws, r, col, value="", bold=False, fg=None, color=None,
             align="left", num_fmt=None, size=9):
        c = ws.cell(row=r, column=col, value=value)
        c.font      = Font(bold=bold, color=color or C["dark"], name="Arial", size=size)
        if fg: c.fill = PatternFill("solid", fgColor=fg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border    = borde
        if num_fmt: c.number_format = num_fmt
        return c

    def head(ws, r, col, value):
        return cell(ws, r, col, value, bold=True, fg=C["azul"], color=C["blanco"], align="center")

    def row_bg(i): return C["gris"] if i % 2 == 0 else C["blanco"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Servicios"
    ws.freeze_panes = "A4"

    NCOLS = 4
    ws.merge_cells(f"A1:{get_column_letter(NCOLS)}1")
    t = ws["A1"]
    t.value     = "Catálogo de Servicios — Fresh Steps"
    t.font      = Font(bold=True, color=C["blanco"], name="Arial", size=13)
    t.fill      = PatternFill("solid", fgColor=C["azul"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(NCOLS)}2")
    subtexto = f"Negocio ID: {id_negocio}" if id_negocio else "Todos los negocios"
    if incluir_eliminados: subtexto += "  |  Incluye eliminados"
    s = ws["A2"]
    s.value     = subtexto
    s.font      = Font(italic=True, color=C["gris_txt"], name="Arial", size=9)
    s.fill      = PatternFill("solid", fgColor=C["azul_cl"])
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    HEADERS = ["Negocio", "Servicio", "Precio ($)", "Estado"]
    for ci, h in enumerate(HEADERS, 1): head(ws, 3, ci, h)
    ws.row_dimensions[3].height = 22

    for i, s in enumerate(servicios):
        r  = i + 4
        bg = row_bg(i)
        cell(ws, r, 1, s.get("negocio", ""), fg=bg)
        cell(ws, r, 2, s.get("nombre",  ""), fg=bg)
        cell(ws, r, 3, float(s.get("precio") or 0), fg=bg, bold=True,
             align="right", num_fmt='"$"#,##0.00')

        activo = s.get("activo", 1)
        ce = ws.cell(row=r, column=4, value="Activo" if activo else "Eliminado")
        ce.font      = Font(bold=True, color=C["verde"] if activo else C["rojo"], name="Arial", size=9)
        ce.fill      = PatternFill("solid", fgColor=C["verde_cl"] if activo else C["rojo_cl"])
        ce.alignment = Alignment(horizontal="center", vertical="center")
        ce.border    = borde
        ws.row_dimensions[r].height = 16

    for ci, w in enumerate([20, 32, 14, 12], 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"servicios_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=nombre,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
import logging
from flask import Blueprint, render_template, jsonify, session, request
from datetime import date

from services.ventas_service import (
    listar_ventas_listas_service,
    registrar_pago_final_service,
    listar_entregas_pendientes_service,
    guardar_venta_service,
)
from ventas import (
    marcar_entregada,
    obtener_venta,
    obtener_detalles_venta,
)

logger = logging.getLogger(__name__)

ventas_bp = Blueprint("ventas", __name__)


@ventas_bp.route("/ventas/guardar", methods=["POST"])
def guardar_venta():
    id_usuario = session.get("id_usuario")
    id_venta, error = guardar_venta_service(request.form, id_usuario)

    if error:
        return jsonify({"ok": False, "error": error}), 400

    return jsonify({"ok": True, "id_venta": id_venta}), 200


@ventas_bp.route("/ventas")
def ventas():
    return render_template("ventas_crear.html")


@ventas_bp.route("/ventas/entregar/<int:id_venta>", methods=["POST"])
def entregar_venta(id_venta):
    id_usuario = session.get("id_usuario")

    try:
        if marcar_entregada(id_venta, id_usuario):
            return jsonify({
                "ok": True,
                "message": "Venta entregada correctamente"
            })
        else:
            return jsonify({
                "ok": False,
                "error": "La venta ya fue entregada o no existe"
            })

    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500


@ventas_bp.route("/ventas/ticket/<int:id_venta>")
def venta_ticket(id_venta):
    venta = obtener_venta(id_venta)

    detalles_dict = obtener_detalles_venta([id_venta])
    detalles = detalles_dict.get(id_venta, [])  

    return render_template(
        "ticket_venta.html",
        venta=venta,
        detalles=detalles
    )

@ventas_bp.route("/ventas/listas")
def ventas_listas():
    id_negocio = request.args.get("id_negocio") or None

    data = listar_ventas_listas_service(id_negocio)

    return render_template(
        "ventas_listas.html",
        **data
    )


@ventas_bp.route("/ventas/pendientes")
def ventas_pendientes():
    try:
        id_negocio = request.args.get("id_negocio", type=int)

        data = listar_entregas_pendientes_service(id_negocio)

        logger.debug("ventas_pendientes data: %s", data)

        return render_template(
            "ventas_pendientes.html",
            **data
        )

    except Exception as e:
        logger.error("ERROR EN ventas_pendientes: %s", e, exc_info=True)
        return str(e), 500


@ventas_bp.route("/ventas/marcar-lista/<int:id_venta>", methods=["POST"])
def marcar_lista(id_venta):
    from ventas import marcar_como_lista

    try:
        if marcar_como_lista(id_venta):
            return jsonify({
                "ok": True,
                "message": "Venta marcada como lista correctamente"
            })
        else:
            return jsonify({
                "ok": False,
                "error": "La venta ya está lista o fue entregada"
            })
    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500
    

@ventas_bp.route("/ventas/detalles/<int:id_venta>")
def detalles_venta(id_venta):
    from ventas import obtener_detalles_venta

    detalles = obtener_detalles_venta([id_venta])  
    return jsonify(detalles.get(id_venta, []))

@ventas_bp.route("/ventas/pago-final", methods=["POST"])
def registrar_pago_final():
    data = request.json
    id_usuario = session.get("id_usuario")

    try:
        ok, mensaje = registrar_pago_final_service(data, id_usuario)

        if not ok:
            return jsonify({
                "ok": False,
                "error": mensaje
            }), 400

        return jsonify({
            "ok": True,
            "message": mensaje
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500
    

@ventas_bp.route("/ventas/eliminar/<int:id_venta>", methods=["POST"])
def eliminar_venta_route(id_venta):
    from services.ventas_service import eliminar_venta_service

    id_usuario = session.get("id_usuario")
    rol = session.get("rol")

    if not id_usuario:
        return jsonify({
            "ok": False,
            "error": "No autenticado"
        }), 401

    if rol != "admin":
        return jsonify({
            "ok": False,
            "error": "No autorizado"
        }), 403

    try:
        ok, mensaje = eliminar_venta_service(id_venta)

        if not ok:
            return jsonify({
                "ok": False,
                "error": mensaje
            }), 400

        return jsonify({
            "ok": True,
            "message": mensaje
        })

    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@ventas_bp.route("/ventas/historial")
def historial_ventas():
    from services.ventas_service import historial_ventas_service

    rol = session.get("rol")
    if rol != "admin":
        from flask import render_template as rt
        return rt("403.html"), 403

    id_negocio   = request.args.get("id_negocio",  type=int)
    fecha_inicio = request.args.get("fecha_inicio") or None
    fecha_fin    = request.args.get("fecha_fin")    or None
    pagina       = request.args.get("pagina", 1,    type=int)

    data = historial_ventas_service(id_negocio, fecha_inicio, fecha_fin, pagina)

    return render_template("historial_ventas.html", **data)



@ventas_bp.route("/ventas/historial/exportar")
def exportar_historial_excel():
    import io
    from datetime import datetime
    from flask import send_file
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from ventas import obtener_historial_ventas, obtener_detalles_venta
    from pagos import obtener_pagos_venta

    if session.get("rol") != "admin":
        from flask import render_template as rt
        return rt("403.html"), 403

    id_negocio   = request.args.get("id_negocio",  type=int)
    fecha_inicio = request.args.get("fecha_inicio") or None
    fecha_fin    = request.args.get("fecha_fin")    or None

    ventas = obtener_historial_ventas(id_negocio, fecha_inicio, fecha_fin, limit=99999, offset=0)

    ids_venta    = [v["id_venta"] for v in ventas]
    detalles_map = obtener_detalles_venta(ids_venta)
    pagos_map    = obtener_pagos_venta(ids_venta)

    C = {
        "azul":        "1E7FD6",
        "azul_claro":  "E6F3FF",
        "azul_med":    "BFDBFE",
        "gris":        "F8FBFF",
        "blanco":      "FFFFFF",
        "verde":       "22C55E",
        "verde_cl":    "DCFCE7",
        "amarillo":    "F59E0B",
        "amarillo_cl": "FEF9C3",
        "rojo":        "E53935",
        "rojo_cl":     "FEE2E2",
        "gris_txt":    "6B7280",
        "dark":        "1F2937",
    }

    borde = Border(
        left=Side(style="thin", color="CFD8E3"),
        right=Side(style="thin", color="CFD8E3"),
        top=Side(style="thin", color="CFD8E3"),
        bottom=Side(style="thin", color="CFD8E3"),
    )
    borde_med = Border(
        left=Side(style="medium", color="1E7FD6"),
        right=Side(style="medium", color="1E7FD6"),
        top=Side(style="medium", color="1E7FD6"),
        bottom=Side(style="medium", color="1E7FD6"),
    )

    def cell(ws, r, col, value="", bold=False, fg=None, color=None,
             align="left", num_fmt=None, wrap=False, italic=False, size=9):
        c = ws.cell(row=r, column=col, value=value)
        c.font      = Font(bold=bold, italic=italic, color=color or C["dark"],
                           name="Arial", size=size)
        if fg:
            c.fill  = PatternFill("solid", fgColor=fg)
        c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        c.border    = borde
        if num_fmt:
            c.number_format = num_fmt
        return c

    def head(ws, r, col, value, size=9):
        c = cell(ws, r, col, value, bold=True, fg=C["azul"], color=C["blanco"],
                 align="center", size=size)
        return c

    def subhead(ws, r, col, value):
        c = cell(ws, r, col, value, bold=True, fg=C["azul_claro"], color=C["azul"],
                 align="left", size=9)
        return c

    def row_bg(row_idx):
        return C["gris"] if row_idx % 2 == 0 else C["blanco"]

    def fmt_dt(dt):
        return dt.strftime("%d/%m/%Y %H:%M") if dt else "—"

    def estado_str(v):
        if v.get("fecha_entrega"): return "Entregada"
        if v.get("fecha_lista"):   return "Lista"
        return "Pendiente"

    COLOR_ESTADO    = {"Entregada": C["verde"],    "Lista": C["amarillo"],    "Pendiente": C["rojo"]}
    COLOR_ESTADO_CL = {"Entregada": C["verde_cl"], "Lista": C["amarillo_cl"], "Pendiente": C["rojo_cl"]}

    def titulo_hoja(ws, texto, ncols, subtexto=""):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value     = texto
        c.font      = Font(bold=True, color=C["blanco"], name="Arial", size=13)
        c.fill      = PatternFill("solid", fgColor=C["azul"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
        s = ws["A2"]
        s.value     = subtexto or "Fresh Steps — Exportación de historial"
        s.font      = Font(italic=True, color=C["gris_txt"], name="Arial", size=9)
        s.fill      = PatternFill("solid", fgColor=C["azul_claro"])
        s.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 16

    filtro_txt = "  ".join(filter(None, [
        f"Negocio ID: {id_negocio}" if id_negocio else "",
        f"Desde: {fecha_inicio}" if fecha_inicio else "",
        f"Hasta: {fecha_fin}" if fecha_fin else "",
    ])) or "Sin filtros — todos los registros"

    wb = Workbook()


    ws1 = wb.active
    ws1.title = "Resumen ventas"
    ws1.freeze_panes = "A4"

    COLS1 = ["# Recibo", "Negocio", "Cliente", "Teléfono",
             "Fecha recibo", "Fecha estimada", "Fecha lista",
             "Fecha entrega", "Total ($)", "Cobrado ($)", "Saldo ($)",
             "Estado", "Registrada por", "Entregada por"]
    titulo_hoja(ws1, "Historial de Ventas", len(COLS1), filtro_txt)
    for ci, h in enumerate(COLS1, 1):
        head(ws1, 3, ci, h)
    ws1.row_dimensions[3].height = 22

    for i, v in enumerate(ventas):
        r       = i + 4
        bg      = row_bg(i)
        estado  = estado_str(v)
        total   = float(v.get("total") or 0)
        cobrado = float(v.get("total_pagado") or 0)
        saldo   = max(total - cobrado, 0)

        cell(ws1, r, 1,  f"#{v['id_venta']}", fg=bg, align="center", bold=True)
        cell(ws1, r, 2,  v.get("negocio", ""), fg=bg)
        cell(ws1, r, 3,  f"{v['nombre']} {v['apellido']}", fg=bg)
        cell(ws1, r, 4,  v.get("telefono") or "—", fg=bg)
        cell(ws1, r, 5,  fmt_dt(v.get("fecha_recibo")), fg=bg)
        cell(ws1, r, 6,  fmt_dt(v.get("fecha_estimada")), fg=bg)
        cell(ws1, r, 7,  fmt_dt(v.get("fecha_lista")), fg=bg)
        cell(ws1, r, 8,  fmt_dt(v.get("fecha_entrega")), fg=bg)
        cell(ws1, r, 9,  total,   fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00')
        cell(ws1, r, 10, cobrado, fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00')
        cell(ws1, r, 11, saldo,   fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00',
             color=C["rojo"] if saldo > 0 else C["verde"])
        # Badge estado
        ce = ws1.cell(row=r, column=12, value=estado)
        ce.font      = Font(bold=True, color=COLOR_ESTADO[estado], name="Arial", size=9)
        ce.fill      = PatternFill("solid", fgColor=COLOR_ESTADO_CL[estado])
        ce.alignment = Alignment(horizontal="center", vertical="center")
        ce.border    = borde
        cell(ws1, r, 13, v.get("usuario_creo") or "—", fg=bg)
        cell(ws1, r, 14, v.get("usuario_entrego") or "—", fg=bg)
        ws1.row_dimensions[r].height = 16

    if ventas:
        tr = len(ventas) + 4
        ws1.merge_cells(f"A{tr}:{get_column_letter(8)}{tr}")
        ct = ws1.cell(row=tr, column=1, value="TOTALES")
        ct.font      = Font(bold=True, color=C["blanco"], name="Arial", size=10)
        ct.fill      = PatternFill("solid", fgColor=C["azul"])
        ct.alignment = Alignment(horizontal="right", vertical="center")
        ct.border    = borde
        fd, ld = 4, len(ventas) + 3
        for col in [9, 10, 11]:
            L = get_column_letter(col)
            tc = ws1.cell(row=tr, column=col, value=f"=SUM({L}{fd}:{L}{ld})")
            tc.font         = Font(bold=True, color=C["blanco"], name="Arial", size=10)
            tc.fill         = PatternFill("solid", fgColor=C["azul"])
            tc.alignment    = Alignment(horizontal="right", vertical="center")
            tc.number_format = '"$"#,##0.00'
            tc.border        = borde
        for col in [12, 13, 14]:
            ec = ws1.cell(row=tr, column=col)
            ec.fill   = PatternFill("solid", fgColor=C["azul"])
            ec.border = borde
        ws1.row_dimensions[tr].height = 20

    for ci, w in enumerate([10,16,26,16,20,20,18,18,13,13,13,13,18,18], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w


    ws2 = wb.create_sheet("Artículos")
    ws2.freeze_panes = "A4"

    COLS2 = ["# Recibo", "Negocio", "Cliente", "Tipo artículo",
             "Descripción", "Material / Notas", "Cantidad",
             "Servicio", "Precio servicio ($)", "Comentario"]
    titulo_hoja(ws2, "Artículos por Venta", len(COLS2), filtro_txt)
    for ci, h in enumerate(COLS2, 1):
        head(ws2, 3, ci, h)
    ws2.row_dimensions[3].height = 22

    r2 = 4
    for v in ventas:
        detalles = detalles_map.get(v["id_venta"], [])
        if not detalles:
            bg = row_bg(r2)
            cell(ws2, r2, 1, f"#{v['id_venta']}", fg=bg, bold=True, align="center")
            cell(ws2, r2, 2, v.get("negocio",""), fg=bg)
            cell(ws2, r2, 3, f"{v['nombre']} {v['apellido']}", fg=bg)
            for ci in range(4, 11):
                cell(ws2, r2, ci, "—", fg=bg, color=C["gris_txt"], align="center")
            ws2.row_dimensions[r2].height = 15
            r2 += 1
            continue

        for det in detalles:
            tipo  = det.get("tipo_articulo", "")
            datos = det.get("datos") or {}
            coment = det.get("comentario") or ""
            servicios = det.get("servicios", [])

            if tipo == "calzado":
                desc  = f"{datos.get('tipo','')} {datos.get('marca','')}".strip()
                mat   = f"Color: {datos.get('color_base','—')}  Material: {datos.get('material','—')}"
                cant  = 1
            elif tipo == "confeccion":
                desc  = f"{datos.get('tipo','')} {datos.get('marca','')}".strip()
                mat   = f"Material: {datos.get('material','—')}"
                cant  = datos.get("cantidad", 1)
            elif tipo == "maquila":
                desc  = datos.get("tipo", "—")
                mat   = f"Precio unitario: ${float(datos.get('precio_unitario') or 0):.2f}"
                cant  = datos.get("cantidad", 1)
            else:
                desc, mat, cant = "—", "—", "—"

            if not servicios:
                bg = row_bg(r2)
                cell(ws2, r2, 1, f"#{v['id_venta']}", fg=bg, bold=True, align="center")
                cell(ws2, r2, 2, v.get("negocio",""), fg=bg)
                cell(ws2, r2, 3, f"{v['nombre']} {v['apellido']}", fg=bg)
                cell(ws2, r2, 4, tipo.capitalize(), fg=bg, align="center")
                cell(ws2, r2, 5, desc, fg=bg)
                cell(ws2, r2, 6, mat,  fg=bg, wrap=True)
                cell(ws2, r2, 7, cant, fg=bg, align="center")
                cell(ws2, r2, 8, "—",  fg=bg, color=C["gris_txt"], align="center")
                cell(ws2, r2, 9, "—",  fg=bg, color=C["gris_txt"], align="center")
                cell(ws2, r2, 10, coment, fg=bg, wrap=True, italic=bool(coment))
                ws2.row_dimensions[r2].height = 15
                r2 += 1
            else:
                for si, svc in enumerate(servicios):
                    bg = row_bg(r2)
                    cell(ws2, r2, 1, f"#{v['id_venta']}" if si == 0 else "", fg=bg, bold=si==0, align="center")
                    cell(ws2, r2, 2, v.get("negocio","") if si == 0 else "", fg=bg)
                    cell(ws2, r2, 3, f"{v['nombre']} {v['apellido']}" if si == 0 else "", fg=bg)
                    cell(ws2, r2, 4, tipo.capitalize() if si == 0 else "", fg=bg, align="center")
                    cell(ws2, r2, 5, desc if si == 0 else "", fg=bg)
                    cell(ws2, r2, 6, mat  if si == 0 else "", fg=bg, wrap=True)
                    cell(ws2, r2, 7, cant if si == 0 else "", fg=bg, align="center")
                    cell(ws2, r2, 8, svc.get("nombre",""), fg=bg)
                    cell(ws2, r2, 9, float(svc.get("precio_aplicado") or 0),
                         fg=bg, align="right", num_fmt='"$"#,##0.00')
                    cell(ws2, r2, 10, coment if si == 0 else "", fg=bg, wrap=True, italic=bool(coment))
                    ws2.row_dimensions[r2].height = 15
                    r2 += 1

    for ci, w in enumerate([10,16,26,14,24,28,10,24,18,24], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w


    ws3 = wb.create_sheet("Pagos")
    ws3.freeze_panes = "A4"

    COLS3 = ["# Recibo", "Negocio", "Cliente", "Tipo pago",
             "Método", "Monto ($)", "Total venta ($)", "Registrada por", "Entregada por"]
    titulo_hoja(ws3, "Pagos por Venta", len(COLS3), filtro_txt)
    for ci, h in enumerate(COLS3, 1):
        head(ws3, 3, ci, h)
    ws3.row_dimensions[3].height = 22

    r3 = 4
    for v in ventas:
        pagos = pagos_map.get(v["id_venta"], [])
        total = float(v.get("total") or 0)

        if not pagos:
            bg = row_bg(r3)
            cell(ws3, r3, 1, f"#{v['id_venta']}", fg=bg, bold=True, align="center")
            cell(ws3, r3, 2, v.get("negocio",""), fg=bg)
            cell(ws3, r3, 3, f"{v['nombre']} {v['apellido']}", fg=bg)
            cell(ws3, r3, 4, "Sin pagos", fg=bg, color=C["gris_txt"], italic=True)
            cell(ws3, r3, 5, "—", fg=bg, color=C["gris_txt"], align="center")
            cell(ws3, r3, 6, "—", fg=bg, color=C["gris_txt"], align="center")
            cell(ws3, r3, 7, total, fg=bg, align="right", num_fmt='"$"#,##0.00')
            cell(ws3, r3, 8, v.get("usuario_creo") or "—", fg=bg)
            cell(ws3, r3, 9, v.get("usuario_entrego") or "—", fg=bg)
            ws3.row_dimensions[r3].height = 15
            r3 += 1
            continue

        for pi, p in enumerate(pagos):
            bg = row_bg(r3)
            cell(ws3, r3, 1, f"#{v['id_venta']}" if pi == 0 else "", fg=bg, bold=pi==0, align="center")
            cell(ws3, r3, 2, v.get("negocio","") if pi == 0 else "", fg=bg)
            cell(ws3, r3, 3, f"{v['nombre']} {v['apellido']}" if pi == 0 else "", fg=bg)
            cell(ws3, r3, 4, (p.get("tipo_pago_venta") or "—").capitalize(), fg=bg)
            cell(ws3, r3, 5, (p.get("tipo_pago") or "—").capitalize(), fg=bg)
            cell(ws3, r3, 6, float(p.get("monto") or 0),
                 fg=bg, bold=True, align="right",
                 color=C["verde"], num_fmt='"$"#,##0.00')
            cell(ws3, r3, 7, total if pi == 0 else "", fg=bg,
                 align="right", num_fmt='"$"#,##0.00')
            cell(ws3, r3, 8, v.get("usuario_creo") or "—" if pi == 0 else "", fg=bg)
            cell(ws3, r3, 9, v.get("usuario_entrego") or "—" if pi == 0 else "", fg=bg)
            ws3.row_dimensions[r3].height = 15
            r3 += 1

    for ci, w in enumerate([10,16,26,16,16,14,14,18,18], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"historial_ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
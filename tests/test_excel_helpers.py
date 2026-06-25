"""
Tests directos sobre services/excel_helpers.py.
Cubre: xl_subhead, fmt_dt (date/str), build_excel_cliente con datos completos,
       _build_ws_articulos (calzado, confeccion, maquila, sin detalles),
       _build_ws_pagos (con y sin pagos), _build_ws_resumen (con pedidos).
"""
from datetime import datetime
from openpyxl import Workbook

from services.excel_helpers import (
    xl_subhead,
    fmt_dt,
    build_excel_cliente,
    _build_ws_resumen,
    _build_ws_articulos,
    _build_ws_pagos,
)


def _pedido(id_venta=1, negocio="Calzado", total=150.0,
            fecha_lista=None, fecha_entrega=None):
    return {
        "id_venta":       id_venta,
        "negocio":        negocio,
        "total":          total,
        "fecha_recibo":   datetime(2026, 1, 1, 10, 0),
        "fecha_estimada": datetime(2026, 1, 15, 10, 0),
        "fecha_lista":    fecha_lista,
        "fecha_entrega":  fecha_entrega,
    }


def _det_calzado(id_venta=1, servicio="Limpieza", precio=150.0):
    return {
        "tipo_articulo": "calzado",
        "datos": {
            "tipo":       "Tenis",
            "marca":      "Nike",
            "color_base": "Blanco",
            "material":   "Piel",
        },
        "comentario": "Test cobertura",
        "servicios":  [{"nombre": servicio, "precio_aplicado": precio}],
    }


def _det_confeccion(id_venta=2):
    return {
        "tipo_articulo": "confeccion",
        "datos": {
            "tipo":     "Camisa",
            "marca":    "Levis",
            "material": "Algodón",
            "cantidad": 3,
        },
        "comentario": "",
        "servicios": [{"nombre": "Bordado", "precio_aplicado": 80.0}],
    }


def _det_maquila(id_venta=3):
    return {
        "tipo_articulo": "maquila",
        "datos": {
            "tipo":            "Mandil",
            "precio_unitario": "20.00",
            "cantidad":        5,
        },
        "comentario": "",
        "servicios": [],
    }


# ---------------------------------------------------------------------------

def test_xl_subhead_retorna_celda():
    """Cubre excel_helpers.py:68 — cuerpo de xl_subhead."""
    wb = Workbook()
    ws = wb.active
    celda = xl_subhead(ws, 1, 1, "Encabezado sub")
    assert celda.value == "Encabezado sub"


def test_fmt_dt_none():
    assert fmt_dt(None) == "—"


def test_fmt_dt_datetime():
    dt = datetime(2026, 6, 24, 10, 30)
    result = fmt_dt(dt)
    assert "2026" in result and "10:30" in result


def test_build_excel_cliente_sin_pedidos():
    """Resumen vacío — ejecuta funciones sin loop."""
    wb = build_excel_cliente([], {}, {}, "Cliente Test", "Sin filtros")
    assert wb is not None
    assert wb.active.title == "Resumen pedidos"


def test_build_ws_resumen_con_pedidos():
    """Cubre _build_ws_resumen líneas 178-198."""
    wb = Workbook()
    ws = wb.active
    pedidos = [_pedido(1, "Calzado", 150.0)]
    pagos_map = {1: [{"tipo_pago_venta": "parcial", "tipo_pago": "efectivo", "monto": 50.0}]}
    _build_ws_resumen(ws, pedidos, pagos_map, "Juan Test", "")
    # Fila de datos en row 4
    assert ws.cell(row=4, column=1).value == "#1"


def test_build_ws_resumen_saldo_cero():
    """Cubre la rama color verde (saldo == 0)."""
    wb = Workbook()
    ws = wb.active
    pedidos = [_pedido(1, total=100.0)]
    pagos_map = {1: [{"tipo_pago_venta": "final", "tipo_pago": "tarjeta", "monto": 100.0}]}
    _build_ws_resumen(ws, pedidos, pagos_map, "Ana", "")
    assert ws.cell(row=4, column=1).value == "#1"


def test_build_ws_articulos_calzado():
    """Cubre _build_ws_articulos con tipo=calzado."""
    wb = Workbook()
    ws = wb.active
    pedidos = [_pedido(1)]
    detalles_map = {1: [_det_calzado(1)]}
    _build_ws_articulos(ws, pedidos, detalles_map, "Cliente", "")
    assert ws.cell(row=4, column=1).value == "#1"


def test_build_ws_articulos_confeccion():
    """Cubre _build_ws_articulos con tipo=confeccion."""
    wb = Workbook()
    ws = wb.active
    pedidos = [_pedido(2, "Confección")]
    detalles_map = {2: [_det_confeccion(2)]}
    _build_ws_articulos(ws, pedidos, detalles_map, "Cliente", "")
    assert ws.cell(row=4, column=3).value == "Confeccion"


def test_build_ws_articulos_maquila():
    """Cubre _build_ws_articulos con tipo=maquila."""
    wb = Workbook()
    ws = wb.active
    pedidos = [_pedido(3, "Maquila")]
    detalles_map = {3: [_det_maquila(3)]}
    _build_ws_articulos(ws, pedidos, detalles_map, "Cliente", "")
    assert ws.cell(row=4, column=3).value == "Maquila"


def test_build_ws_articulos_sin_detalles():
    """Cubre _build_ws_articulos lines 211-217 — venta sin detalles."""
    wb = Workbook()
    ws = wb.active
    pedidos = [_pedido(99)]
    detalles_map = {}  # sin detalles → rama if not detalles
    _build_ws_articulos(ws, pedidos, detalles_map, "Cliente", "")
    assert ws.cell(row=4, column=1).value == "#99"


def test_build_ws_pagos_con_pagos():
    """Cubre _build_ws_pagos con pagos."""
    wb = Workbook()
    ws = wb.active
    pedidos = [_pedido(1)]
    pagos_map = {1: [
        {"tipo_pago_venta": "parcial", "tipo_pago": "efectivo", "monto": 50.0},
        {"tipo_pago_venta": "final",   "tipo_pago": "tarjeta",  "monto": 100.0},
    ]}
    _build_ws_pagos(ws, pedidos, pagos_map, "Cliente", "")
    assert ws.cell(row=4, column=1).value == "#1"


def test_build_ws_pagos_sin_pagos():
    """Cubre _build_ws_pagos lines 262-270 — venta sin pagos."""
    wb = Workbook()
    ws = wb.active
    pedidos = [_pedido(1)]
    pagos_map = {}  # sin pagos
    _build_ws_pagos(ws, pedidos, pagos_map, "Cliente", "")
    # "Sin pagos" debería aparecer
    assert ws.cell(row=4, column=3).value == "Sin pagos"


def test_build_excel_cliente_completo():
    """Integración de los tres sheets con datos reales."""
    pedidos = [
        _pedido(1, "Calzado", 150.0),
        _pedido(2, "Confección", 200.0, fecha_lista=datetime(2026, 2, 1)),
        _pedido(3, "Maquila", 100.0, fecha_lista=datetime(2026, 2, 5), fecha_entrega=datetime(2026, 2, 10)),
    ]
    detalles_map = {
        1: [_det_calzado(1)],
        2: [_det_confeccion(2)],
        3: [_det_maquila(3)],
    }
    pagos_map = {
        1: [{"tipo_pago_venta": "parcial", "tipo_pago": "efectivo", "monto": 75.0}],
        3: [{"tipo_pago_venta": "final",   "tipo_pago": "tarjeta",  "monto": 100.0}],
    }
    wb = build_excel_cliente(pedidos, detalles_map, pagos_map, "María García", "Todos")
    sheets = [s.title for s in wb.worksheets]
    assert "Resumen pedidos" in sheets
    assert "Artículos" in sheets
    assert "Pagos" in sheets

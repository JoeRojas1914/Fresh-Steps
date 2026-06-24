"""
Tests para services/excel_ventas_service.py
Verifica que exportar_historial_service() devuelva un Workbook válido con las
3 hojas esperadas, tanto con datos como sin ellos.
"""
import pytest
from openpyxl import Workbook
from services.excel_ventas_service import exportar_historial_service

HOJAS_ESPERADAS = ["Resumen ventas", "Artículos", "Pagos"]


def test_exportar_sin_filtros_devuelve_workbook():
    wb = exportar_historial_service(None, None, None)
    assert isinstance(wb, Workbook)


def test_exportar_tiene_tres_hojas():
    wb = exportar_historial_service(None, None, None)
    assert wb.sheetnames == HOJAS_ESPERADAS


def test_exportar_con_rango_de_fechas():
    wb = exportar_historial_service(None, "2026-01-01", "2026-06-30")
    assert isinstance(wb, Workbook)
    assert wb.sheetnames == HOJAS_ESPERADAS


def test_exportar_filtrado_por_negocio():
    wb = exportar_historial_service(1, "2026-01-01", "2026-06-30")
    assert isinstance(wb, Workbook)


def test_exportar_con_tipo_fecha_lista():
    wb = exportar_historial_service(None, "2026-01-01", "2026-06-30", tipo_fecha="fecha_lista")
    assert isinstance(wb, Workbook)


def test_exportar_con_tipo_fecha_entrega():
    wb = exportar_historial_service(None, "2026-01-01", "2026-06-30", tipo_fecha="fecha_entrega")
    assert isinstance(wb, Workbook)


def test_exportar_resumen_tiene_headers():
    wb = exportar_historial_service(None, None, None)
    ws = wb["Resumen ventas"]
    # Fila 3 contiene los headers (filas 1-2 son el título)
    headers = [ws.cell(row=3, column=c).value for c in range(1, 15)]
    assert "# Recibo" in headers
    assert "Total ($)" in headers
    assert "Estado" in headers


def test_exportar_articulos_tiene_headers():
    wb = exportar_historial_service(None, None, None)
    ws = wb["Artículos"]
    headers = [ws.cell(row=3, column=c).value for c in range(1, 11)]
    assert "# Recibo" in headers
    assert "Tipo artículo" in headers


def test_exportar_pagos_tiene_headers():
    wb = exportar_historial_service(None, None, None)
    ws = wb["Pagos"]
    headers = [ws.cell(row=3, column=c).value for c in range(1, 10)]
    assert "# Recibo" in headers
    assert "Monto ($)" in headers

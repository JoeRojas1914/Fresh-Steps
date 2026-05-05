import io
from datetime import datetime, date
from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


C = {
    "azul":        "1E7FD6",
    "azul_cl":     "E6F3FF",
    "azul_med":    "BFDBFE",
    "blanco":      "FFFFFF",
    "gris":        "F8FBFF",
    "verde":       "22C55E",
    "verde_cl":    "DCFCE7",
    "amarillo":    "F59E0B",
    "amarillo_cl": "FEF9C3",
    "rojo":        "E53935",
    "rojo_cl":     "FEE2E2",
    "gris_txt":    "6B7280",
    "dark":        "1F2937",
}

COLOR_ESTADO    = {"Entregada": C["verde"],    "Lista": C["amarillo"],    "Pendiente": C["rojo"]}
COLOR_ESTADO_CL = {"Entregada": C["verde_cl"], "Lista": C["amarillo_cl"], "Pendiente": C["rojo_cl"]}

BORDE = Border(
    left=Side(style="thin", color="CFD8E3"),
    right=Side(style="thin", color="CFD8E3"),
    top=Side(style="thin", color="CFD8E3"),
    bottom=Side(style="thin", color="CFD8E3"),
)

BORDE_MED = Border(
    left=Side(style="medium", color="1E7FD6"),
    right=Side(style="medium", color="1E7FD6"),
    top=Side(style="medium", color="1E7FD6"),
    bottom=Side(style="medium", color="1E7FD6"),
)


def xl_cell(ws, r, col, value="", bold=False, fg=None, color=None,
            align="left", num_fmt=None, italic=False, size=9, wrap=False):
    """Escribe y estiliza una celda."""
    c = ws.cell(row=r, column=col, value=value)
    c.font = Font(
        bold=bold, italic=italic,
        color=color or C["dark"],
        name="Arial", size=size
    )
    if fg:
        c.fill = PatternFill("solid", fgColor=fg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    c.border = BORDE
    if num_fmt:
        c.number_format = num_fmt
    return c


def xl_head(ws, r, col, value, size=9):
    return xl_cell(ws, r, col, value,
                   bold=True, fg=C["azul"], color=C["blanco"],
                   align="center", size=size)


def xl_subhead(ws, r, col, value):
    return xl_cell(ws, r, col, value,
                   bold=True, fg=C["azul_cl"], color=C["azul"], align="left")


def xl_row_bg(i):
    return C["gris"] if i % 2 == 0 else C["blanco"]


def xl_titulo_hoja(ws, texto, ncols, subtexto=""):
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    t = ws["A1"]
    t.value     = texto
    t.font      = Font(bold=True, color=C["blanco"], name="Arial", size=13)
    t.fill      = PatternFill("solid", fgColor=C["azul"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(ncols)}2")
    s = ws["A2"]
    s.value     = subtexto or "Fresh Steps"
    s.font      = Font(italic=True, color=C["gris_txt"], name="Arial", size=9)
    s.fill      = PatternFill("solid", fgColor=C["azul_cl"])
    s.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16


def xl_fila_headers(ws, headers, fila=3, altura=22):
    for ci, h in enumerate(headers, 1):
        xl_head(ws, fila, ci, h)
    ws.row_dimensions[fila].height = altura


def xl_fila_totales(ws, fila, ncols, cols_suma):
    fd = 4
    ld = fila - 1

    ws.merge_cells(f"A{fila}:{get_column_letter(cols_suma[0] - 1)}{fila}")
    ct = ws.cell(row=fila, column=1, value="TOTALES")
    ct.font      = Font(bold=True, color=C["blanco"], name="Arial", size=10)
    ct.fill      = PatternFill("solid", fgColor=C["azul"])
    ct.alignment = Alignment(horizontal="right", vertical="center")
    ct.border    = BORDE

    for col in cols_suma:
        L  = get_column_letter(col)
        tc = ws.cell(row=fila, column=col, value=f"=SUM({L}{fd}:{L}{ld})")
        tc.font         = Font(bold=True, color=C["blanco"], name="Arial", size=10)
        tc.fill         = PatternFill("solid", fgColor=C["azul"])
        tc.alignment    = Alignment(horizontal="right", vertical="center")
        tc.number_format = '"$"#,##0.00'
        tc.border        = BORDE

    for col in range(1, ncols + 1):
        if col not in cols_suma and col != 1:
            ec = ws.cell(row=fila, column=col)
            if not ec.fill or ec.fill.fgColor.rgb == "00000000":
                ec.fill   = PatternFill("solid", fgColor=C["azul"])
                ec.border = BORDE

    ws.row_dimensions[fila].height = 20


def xl_col_widths(ws, widths):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def fmt_dt(dt):
    if not dt:
        return "—"
    if isinstance(dt, datetime):
        return dt.strftime("%d/%m/%Y %H:%M")
    if isinstance(dt, date):
        return dt.strftime("%d/%m/%Y")
    return str(dt)


def estado_venta(v):
    if v.get("fecha_entrega"):
        return "Entregada"
    if v.get("fecha_lista"):
        return "Lista"
    return "Pendiente"


def xl_badge_estado(ws, r, col, estado, borde=None):
    ce = ws.cell(row=r, column=col, value=estado)
    ce.font      = Font(bold=True, color=COLOR_ESTADO[estado], name="Arial", size=9)
    ce.fill      = PatternFill("solid", fgColor=COLOR_ESTADO_CL[estado])
    ce.alignment = Alignment(horizontal="center", vertical="center")
    ce.border    = borde or BORDE


def xl_badge_activo(ws, r, col, activo, borde=None):
    valor = "Activo" if activo else "Eliminado"
    color = C["verde"] if activo else C["rojo"]
    fondo = C["verde_cl"] if activo else C["rojo_cl"]
    ce = ws.cell(row=r, column=col, value=valor)
    ce.font      = Font(bold=True, color=color, name="Arial", size=9)
    ce.fill      = PatternFill("solid", fgColor=fondo)
    ce.alignment = Alignment(horizontal="center", vertical="center")
    ce.border    = borde or BORDE


def send_excel(wb, nombre_base):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nombre = f"{nombre_base}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
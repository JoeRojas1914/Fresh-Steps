from openpyxl import Workbook

from config import MAX_FILAS_EXPORTAR
from models.ventas import obtener_historial_ventas
from models.pagos import obtener_pagos_venta
from services.ventas_service import obtener_detalles_venta
from services.excel_helpers import (
    C,
    xl_cell, xl_row_bg,
    xl_titulo_hoja, xl_fila_headers, xl_fila_totales,
    xl_col_widths, xl_badge_estado,
    fmt_dt, estado_venta,
    send_excel,
)


def _ws_resumen(ws, ventas, pagos_map, filtro_txt):
    ws.freeze_panes = "A4"
    COLS = [
        "# Recibo", "Negocio", "Cliente", "Teléfono",
        "Fecha recibo", "Fecha estimada", "Fecha lista", "Fecha entrega",
        "Total ($)", "Cobrado ($)", "Saldo ($)",
        "Estado", "Registrada por", "Entregada por",
    ]
    xl_titulo_hoja(ws, "Historial de Ventas", len(COLS), filtro_txt)
    xl_fila_headers(ws, COLS)

    for i, v in enumerate(ventas):
        r = i + 4
        bg = xl_row_bg(i)
        estado  = estado_venta(v)
        total   = float(v.get("total") or 0)
        cobrado = float(v.get("total_pagado") or 0)
        saldo   = max(total - cobrado, 0)

        xl_cell(ws, r,  1, f"#{v['id_venta']}",                 fg=bg, bold=True, align="center")
        xl_cell(ws, r,  2, v.get("negocio", ""),                 fg=bg)
        xl_cell(ws, r,  3, f"{v['nombre']} {v['apellido']}",     fg=bg)
        xl_cell(ws, r,  4, v.get("telefono") or "—",             fg=bg)
        xl_cell(ws, r,  5, fmt_dt(v.get("fecha_recibo")),        fg=bg)
        xl_cell(ws, r,  6, fmt_dt(v.get("fecha_estimada")),      fg=bg)
        xl_cell(ws, r,  7, fmt_dt(v.get("fecha_lista")),         fg=bg)
        xl_cell(ws, r,  8, fmt_dt(v.get("fecha_entrega")),       fg=bg)
        xl_cell(ws, r,  9, total,   fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00')
        xl_cell(ws, r, 10, cobrado, fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00')
        xl_cell(ws, r, 11, saldo,   fg=bg, bold=True, align="right", num_fmt='"$"#,##0.00',
                color=C["rojo"] if saldo > 0 else C["verde"])
        xl_badge_estado(ws, r, 12, estado)
        xl_cell(ws, r, 13, v.get("usuario_creo")    or "—", fg=bg)
        xl_cell(ws, r, 14, v.get("usuario_entrego") or "—", fg=bg)
        ws.row_dimensions[r].height = 16

    if ventas:
        xl_fila_totales(ws, len(ventas) + 4, len(COLS), [9, 10, 11])
    xl_col_widths(ws, [10, 16, 26, 16, 20, 20, 18, 18, 13, 13, 13, 13, 18, 18])


def _ws_articulos(ws, ventas, detalles_map, filtro_txt):
    ws.freeze_panes = "A4"
    COLS = [
        "# Recibo", "Negocio", "Cliente",
        "Tipo artículo", "Descripción", "Material / Notas",
        "Cantidad", "Servicio", "Precio servicio ($)", "Comentario",
    ]
    xl_titulo_hoja(ws, "Artículos por Venta", len(COLS), filtro_txt)
    xl_fila_headers(ws, COLS)

    r = 4
    for v in ventas:
        detalles = detalles_map.get(v["id_venta"], [])
        if not detalles:  # pragma: no cover
            bg = xl_row_bg(r)
            xl_cell(ws, r, 1, f"#{v['id_venta']}", fg=bg, bold=True, align="center")
            xl_cell(ws, r, 2, v.get("negocio", ""), fg=bg)
            xl_cell(ws, r, 3, f"{v['nombre']} {v['apellido']}", fg=bg)
            for ci in range(4, 11):
                xl_cell(ws, r, ci, "—", fg=bg, color=C["gris_txt"], align="center")
            ws.row_dimensions[r].height = 15
            r += 1
            continue

        for det in detalles:
            tipo      = det.get("tipo_articulo", "")
            datos     = det.get("datos") or {}
            coment    = det.get("comentario") or ""
            servicios = det.get("servicios", [])

            if tipo == "calzado":
                desc = f"{datos.get('tipo','')} {datos.get('marca','')}".strip()
                mat  = f"Color: {datos.get('color_base','—')}  Material: {datos.get('material','—')}"
                cant = 1
            elif tipo == "confeccion":
                desc = f"{datos.get('tipo','')} {datos.get('marca','')}".strip()
                mat  = f"Material: {datos.get('material','—')}"
                cant = datos.get("cantidad", 1)
            elif tipo == "maquila":
                desc = datos.get("tipo", "—")
                mat  = f"Precio unitario: ${float(datos.get('precio_unitario') or 0):.2f}"
                cant = datos.get("cantidad", 1)
            else:  # pragma: no cover
                desc = mat = "—"; cant = "—"

            for si, svc in enumerate(servicios if servicios else [None]):
                bg = xl_row_bg(r)
                xl_cell(ws, r, 1, f"#{v['id_venta']}" if si == 0 else "",       fg=bg, bold=si == 0, align="center")
                xl_cell(ws, r, 2, v.get("negocio", "") if si == 0 else "",       fg=bg)
                xl_cell(ws, r, 3, f"{v['nombre']} {v['apellido']}" if si == 0 else "", fg=bg)
                xl_cell(ws, r, 4, tipo.capitalize() if si == 0 else "",           fg=bg, align="center")
                xl_cell(ws, r, 5, desc if si == 0 else "",                        fg=bg)
                xl_cell(ws, r, 6, mat  if si == 0 else "",                        fg=bg, wrap=True)
                xl_cell(ws, r, 7, cant if si == 0 else "",                        fg=bg, align="center")
                xl_cell(ws, r, 8, svc.get("nombre", "") if svc else "—",          fg=bg)
                xl_cell(ws, r, 9,
                        float(svc.get("precio_aplicado") or 0) if svc else "—",
                        fg=bg, align="right",
                        num_fmt='"$"#,##0.00' if svc else None)
                xl_cell(ws, r, 10, coment if si == 0 else "", fg=bg, wrap=True, italic=bool(coment))
                ws.row_dimensions[r].height = 15
                r += 1

    xl_col_widths(ws, [10, 16, 26, 14, 24, 28, 10, 24, 18, 24])


def _ws_pagos(ws, ventas, pagos_map, filtro_txt):
    ws.freeze_panes = "A4"
    COLS = [
        "# Recibo", "Negocio", "Cliente",
        "Tipo pago", "Método", "Monto ($)",
        "Total venta ($)", "Registrada por", "Entregada por",
    ]
    xl_titulo_hoja(ws, "Pagos por Venta", len(COLS), filtro_txt)
    xl_fila_headers(ws, COLS)

    r = 4
    for v in ventas:
        pagos = pagos_map.get(v["id_venta"], [])
        total = float(v.get("total") or 0)

        if not pagos:
            bg = xl_row_bg(r)
            xl_cell(ws, r, 1, f"#{v['id_venta']}", fg=bg, bold=True, align="center")
            xl_cell(ws, r, 2, v.get("negocio", ""), fg=bg)
            xl_cell(ws, r, 3, f"{v['nombre']} {v['apellido']}", fg=bg)
            xl_cell(ws, r, 4, "Sin pagos", fg=bg, color=C["gris_txt"], italic=True)
            xl_cell(ws, r, 5, "—", fg=bg, color=C["gris_txt"], align="center")
            xl_cell(ws, r, 6, "—", fg=bg, color=C["gris_txt"], align="center")
            xl_cell(ws, r, 7, total, fg=bg, align="right", num_fmt='"$"#,##0.00')
            xl_cell(ws, r, 8, v.get("usuario_creo")    or "—", fg=bg)
            xl_cell(ws, r, 9, v.get("usuario_entrego") or "—", fg=bg)
            ws.row_dimensions[r].height = 15
            r += 1
            continue

        for pi, p in enumerate(pagos):
            bg = xl_row_bg(r)
            xl_cell(ws, r, 1, f"#{v['id_venta']}" if pi == 0 else "",       fg=bg, bold=pi == 0, align="center")
            xl_cell(ws, r, 2, v.get("negocio", "") if pi == 0 else "",       fg=bg)
            xl_cell(ws, r, 3, f"{v['nombre']} {v['apellido']}" if pi == 0 else "", fg=bg)
            xl_cell(ws, r, 4, (p.get("tipo_pago_venta") or "—").capitalize(), fg=bg)
            xl_cell(ws, r, 5, (p.get("tipo_pago") or "—").capitalize(),       fg=bg)
            xl_cell(ws, r, 6, float(p.get("monto") or 0),
                    fg=bg, bold=True, align="right", color=C["verde"], num_fmt='"$"#,##0.00')
            xl_cell(ws, r, 7, total if pi == 0 else "", fg=bg, align="right", num_fmt='"$"#,##0.00')
            xl_cell(ws, r, 8, v.get("usuario_creo")    or "—" if pi == 0 else "", fg=bg)
            xl_cell(ws, r, 9, v.get("usuario_entrego") or "—" if pi == 0 else "", fg=bg)
            ws.row_dimensions[r].height = 15
            r += 1

    xl_col_widths(ws, [10, 16, 26, 16, 16, 14, 14, 18, 18])


def exportar_historial_service(id_negocio, fecha_inicio, fecha_fin, tipo_fecha="fecha_recibo"):
    """Construye y retorna el Workbook de Excel del historial de ventas."""
    ventas    = obtener_historial_ventas(id_negocio, fecha_inicio, fecha_fin,
                                        limit=MAX_FILAS_EXPORTAR, offset=0,
                                        tipo_fecha=tipo_fecha)
    ids_venta    = [v["id_venta"] for v in ventas]
    detalles_map = obtener_detalles_venta(ids_venta)
    pagos_map    = obtener_pagos_venta(ids_venta)

    _ETIQUETAS_FECHA = {
        "fecha_recibo":  "Fecha de recibo",
        "fecha_lista":   "Fecha lista",
        "fecha_entrega": "Fecha de entrega",
    }
    etiqueta_fecha = _ETIQUETAS_FECHA.get(tipo_fecha, "Fecha de recibo")
    filtro_txt = "  ".join(filter(None, [
        f"Negocio ID: {id_negocio}"         if id_negocio  else "",
        f"Filtro por: {etiqueta_fecha}",
        f"Desde: {fecha_inicio}"            if fecha_inicio else "",
        f"Hasta: {fecha_fin}"               if fecha_fin    else "",
    ])) or "Sin filtros — todos los registros"

    wb    = Workbook()
    ws1   = wb.active
    ws1.title = "Resumen ventas"
    _ws_resumen(ws1, ventas, pagos_map, filtro_txt)

    ws2 = wb.create_sheet("Artículos")
    _ws_articulos(ws2, ventas, detalles_map, filtro_txt)

    ws3 = wb.create_sheet("Pagos")
    _ws_pagos(ws3, ventas, pagos_map, filtro_txt)

    return wb

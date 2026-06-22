import logging
from datetime import date, datetime, timedelta
from models.clientes import contar_clientes
from models.servicios import contar_servicios
from models.negocio import obtener_negocios

from models.estadisticas import (
    contar_ventas_por_semana,
    obtener_gastos_por_semana_y_proveedor,
    obtener_gastos_por_semana_y_categoria,
    obtener_total_gastos,
    obtener_total_ingresos,
    obtener_unidades_por_semana,
    obtener_ingresos_por_semana,
    obtener_ventas_con_y_sin_prepago,
    obtener_uso_servicios,
    obtener_ventas_por_dia,
    obtener_ticket_promedio,
    obtener_saldo_por_cobrar,
    obtener_top_clientes,
    obtener_tiempo_promedio_entrega,
    obtener_ingresos_por_negocio,
    obtener_metodos_pago,
    obtener_hora_pico_recepcion,
    obtener_hora_pico_entrega,
    obtener_clientes_unicos,
    obtener_clientes_nuevos,
    obtener_tasa_retorno,
    obtener_gasto_promedio_cliente,
    contar_ventas_por_hora,
    obtener_ingresos_por_hora,
    obtener_unidades_por_hora,
    contar_ventas_por_dia_rango,
    obtener_ingresos_por_dia_rango,
    obtener_unidades_por_dia_rango,
)

MAX_DIAS_RANGO = 186
_COLS_FECHA_ESTADISTICAS = frozenset({"fecha_recibo", "fecha_entrega"})

_log = logging.getLogger(__name__)


def dashboard_page_data_service():
    hoy          = date.today()
    fecha_inicio = hoy.replace(day=1)
    fecha_fin    = hoy
    return {
        "total_clientes":  contar_clientes(),
        "total_servicios": contar_servicios(),
        "negocios":        obtener_negocios(),
        "fecha_inicio":    fecha_inicio.isoformat(),
        "fecha_fin":       fecha_fin.isoformat(),
    }


def _periodo_anterior(inicio: date, fin: date):
    """Devuelve (inicio, fin) del período inmediatamente anterior de igual duración."""
    duracion   = (fin - inicio).days + 1
    fin_ant    = inicio - timedelta(days=1)
    inicio_ant = fin_ant - timedelta(days=duracion - 1)
    return inicio_ant, fin_ant


def _pct_cambio(actual, anterior) -> float | None:
    if anterior == 0:
        return None
    return round((actual - anterior) / anterior * 100, 1)


def _calcular_kpis(inicio: date, fin: date, id_negocio, col: str = "fecha_recibo") -> dict:
    total_ingresos              = obtener_total_ingresos(inicio, fin, id_negocio)
    total_gastos                = obtener_total_gastos(inicio,   fin, id_negocio)
    ticket_promedio, num_ventas = obtener_ticket_promedio(inicio, fin, id_negocio, col)
    saldo_por_cobrar            = obtener_saldo_por_cobrar(inicio, fin, id_negocio, col)
    return {
        "ingresos":         total_ingresos,
        "gastos":           total_gastos,
        "ganancia":         total_ingresos - total_gastos,
        "ticket_promedio":  ticket_promedio,
        "num_ventas":       num_ventas,
        "total_ventas":     num_ventas,
        "saldo_por_cobrar": saldo_por_cobrar,
    }


def _calcular_comparativa(kpis: dict, inicio: date, fin: date, id_negocio, col: str = "fecha_recibo") -> tuple[dict, dict]:
    inicio_ant, fin_ant    = _periodo_anterior(inicio, fin)
    ingresos_ant           = obtener_total_ingresos(inicio_ant, fin_ant, id_negocio)
    gastos_ant             = obtener_total_gastos(inicio_ant,   fin_ant, id_negocio)
    ticket_ant, ventas_ant = obtener_ticket_promedio(inicio_ant, fin_ant, id_negocio, col)
    saldo_ant              = obtener_saldo_por_cobrar(inicio_ant, fin_ant, id_negocio, col)
    ganancia_ant           = ingresos_ant - gastos_ant

    pcts = {
        "ingresos_pct": _pct_cambio(kpis["ingresos"],         ingresos_ant),
        "gastos_pct":   _pct_cambio(kpis["gastos"],           gastos_ant),
        "ganancia_pct": _pct_cambio(kpis["ganancia"],         ganancia_ant),
        "ticket_pct":   _pct_cambio(kpis["ticket_promedio"],  ticket_ant),
        "ventas_pct":   _pct_cambio(kpis["num_ventas"],       ventas_ant),
        "saldo_pct":    _pct_cambio(kpis["saldo_por_cobrar"], saldo_ant),
    }
    periodo_anterior_str = {"inicio": inicio_ant.isoformat(), "fin": fin_ant.isoformat()}
    return pcts, periodo_anterior_str


def _cargar_series(inicio: date, fin: date, id_negocio, granularidad: str, col: str = "fecha_recibo", agrupacion_gastos: str = "proveedor") -> dict:
    if granularidad == "hora":
        ventas_semanales   = contar_ventas_por_hora(inicio, fin, id_negocio, col)
        ingresos_semanales = obtener_ingresos_por_hora(inicio, fin, id_negocio)
        unidades_semanales = obtener_unidades_por_hora(inicio, fin, id_negocio, col)
    elif granularidad == "dia":
        ventas_semanales   = contar_ventas_por_dia_rango(inicio, fin, id_negocio, col)
        ingresos_semanales = obtener_ingresos_por_dia_rango(inicio, fin, id_negocio)
        unidades_semanales = obtener_unidades_por_dia_rango(inicio, fin, id_negocio, col)
    else:
        ventas_semanales   = contar_ventas_por_semana(inicio, fin, id_negocio, col)
        ingresos_semanales = obtener_ingresos_por_semana(inicio, fin, id_negocio)
        unidades_semanales = obtener_unidades_por_semana(inicio, fin, id_negocio, col)

    if agrupacion_gastos == "categoria":
        gastos_semanales = obtener_gastos_por_semana_y_categoria(inicio, fin, id_negocio)
    else:
        gastos_semanales = obtener_gastos_por_semana_y_proveedor(inicio, fin, id_negocio)
    ventas_prepago     = obtener_ventas_con_y_sin_prepago(inicio, fin, id_negocio, col)
    uso_servicios      = obtener_uso_servicios(inicio, fin, id_negocio, col)
    ventas_por_dia     = obtener_ventas_por_dia(inicio, fin, id_negocio, col)
    top_clientes       = obtener_top_clientes(inicio, fin, id_negocio, col=col)
    tiempo_entrega     = obtener_tiempo_promedio_entrega(inicio, fin, id_negocio)
    ingresos_x_negocio = obtener_ingresos_por_negocio(inicio, fin, col) if id_negocio == "all" else []
    metodos_pago       = obtener_metodos_pago(inicio, fin, id_negocio)
    hora_recepcion     = obtener_hora_pico_recepcion(inicio, fin, id_negocio)
    hora_entrega       = obtener_hora_pico_entrega(inicio, fin, id_negocio)
    clientes_unicos    = obtener_clientes_unicos(inicio, fin, id_negocio, col)

    try:
        clientes_nuevos = obtener_clientes_nuevos(inicio, fin, id_negocio, col)
    except Exception as e:
        _log.error("clientes_nuevos: %s", e)
        clientes_nuevos = 0

    try:
        tasa_retorno = obtener_tasa_retorno(inicio, fin, id_negocio, col)
    except Exception as e:
        _log.error("tasa_retorno: %s", e)
        tasa_retorno = {"total": 0, "recurrentes": 0, "tasa": 0}

    try:
        gasto_prom_cliente = obtener_gasto_promedio_cliente(inicio, fin, id_negocio, col)
    except Exception as e:
        _log.error("gasto_prom_cliente: %s", e)
        gasto_prom_cliente = 0.0

    return {
        "ventas_semanales":   ventas_semanales,
        "gastos_semanales":   gastos_semanales,
        "ingresos_semanales": ingresos_semanales,
        "unidades_semanales": unidades_semanales,
        "ventas_prepago":     ventas_prepago,
        "uso_servicios":      uso_servicios,
        "ventas_por_dia":     ventas_por_dia,
        "top_clientes":       top_clientes,
        "tiempo_entrega":     tiempo_entrega,
        "ingresos_x_negocio": ingresos_x_negocio,
        "metodos_pago":       metodos_pago,
        "hora_recepcion":     hora_recepcion,
        "hora_entrega":       hora_entrega,
        "clientes_unicos":    clientes_unicos,
        "clientes_nuevos":    clientes_nuevos,
        "tasa_retorno":       tasa_retorno,
        "gasto_prom_cliente": gasto_prom_cliente,
    }


def exportar_estadisticas_service(args):
    inicio_str = args.get("inicio")
    fin_str    = args.get("fin")
    id_negocio = args.get("id_negocio", "all")
    col        = args.get("tipo_fecha", "fecha_recibo")
    if col not in _COLS_FECHA_ESTADISTICAS:
        col = "fecha_recibo"

    if not inicio_str or not fin_str:
        return None, "Faltan fechas"
    try:
        inicio = datetime.strptime(inicio_str, "%Y-%m-%d").date()
        fin    = datetime.strptime(fin_str,    "%Y-%m-%d").date()
    except ValueError:
        return None, "Formato de fecha inválido"
    if fin < inicio:
        return None, "La fecha fin no puede ser menor a inicio"
    if (fin - inicio).days > MAX_DIAS_RANGO:
        return None, f"El rango máximo permitido es {MAX_DIAS_RANGO} días"

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from services.excel_helpers import (
        C, xl_cell, xl_titulo_hoja, xl_fila_headers, xl_col_widths, send_excel
    )

    kpis   = _calcular_kpis(inicio, fin, id_negocio, col)
    series = _cargar_series(inicio, fin, id_negocio, "dia" if (fin - inicio).days <= 31 else "semana", col)

    negocio_label = {"1": "Calzado", "2": "Confección", "3": "Maquila"}.get(str(id_negocio), "Todos los negocios")
    subtitulo = f"{inicio_str} al {fin_str} — {negocio_label}"

    wb = Workbook()

    # ── Hoja 1: KPIs ──
    ws1 = wb.active
    ws1.title = "Resumen"
    xl_titulo_hoja(ws1, "Estadísticas Fresh Steps", 2, subtitulo)
    xl_fila_headers(ws1, ["Indicador", "Valor"], fila=3)

    kpi_rows = [
        ("Ingresos",          kpis["ingresos"],         '#,##0.00'),
        ("Gastos",            kpis["gastos"],            '#,##0.00'),
        ("Ganancia",          kpis["ganancia"],          '#,##0.00'),
        ("Total de ventas",   kpis["num_ventas"],        '#,##0'),
        ("Ticket promedio",   kpis["ticket_promedio"],   '#,##0.00'),
        ("Saldo por cobrar",  kpis["saldo_por_cobrar"],  '#,##0.00'),
    ]
    for i, (label, value, fmt) in enumerate(kpi_rows, 4):
        bg = C["gris"] if i % 2 == 0 else C["blanco"]
        xl_cell(ws1, i, 1, label, fg=bg)
        xl_cell(ws1, i, 2, value, fg=bg, align="right", num_fmt=fmt)
    xl_col_widths(ws1, {1: 24, 2: 18})

    # ── Hoja 2: Series ──
    ws2 = wb.create_sheet("Series por período")
    labels    = [r["label"] for r in series.get("ventas_semanales", [])]
    ventas    = [r["total"] for r in series.get("ventas_semanales", [])]
    ingresos  = [r["total"] for r in series.get("ingresos_semanales", [])]
    unidades  = [r["total"] for r in series.get("unidades_semanales", [])]

    xl_titulo_hoja(ws2, "Series por período", 4, subtitulo)
    xl_fila_headers(ws2, ["Período", "Ventas", "Ingresos ($)", "Unidades"], fila=3)
    for i, lbl in enumerate(labels):
        row = i + 4
        bg  = C["gris"] if i % 2 == 0 else C["blanco"]
        xl_cell(ws2, row, 1, lbl,             fg=bg)
        xl_cell(ws2, row, 2, ventas[i],       fg=bg, align="right", num_fmt='#,##0')
        xl_cell(ws2, row, 3, ingresos[i],     fg=bg, align="right", num_fmt='#,##0.00')
        xl_cell(ws2, row, 4, unidades[i],     fg=bg, align="right", num_fmt='#,##0')
    xl_col_widths(ws2, {1: 20, 2: 12, 3: 16, 4: 12})

    return send_excel(wb, "estadisticas"), None


def dashboard_api_service(args):
    inicio_str = args.get("inicio")
    fin_str    = args.get("fin")
    id_negocio = args.get("id_negocio", "all")
    col        = args.get("tipo_fecha", "fecha_recibo")
    if col not in _COLS_FECHA_ESTADISTICAS:
        col = "fecha_recibo"

    if not inicio_str or not fin_str:
        return None, "Faltan fechas"

    try:
        inicio = datetime.strptime(inicio_str, "%Y-%m-%d").date()
        fin    = datetime.strptime(fin_str,    "%Y-%m-%d").date()
    except ValueError:
        return None, "Formato de fecha inválido (usa YYYY-MM-DD)"

    if fin < inicio:
        return None, "La fecha fin no puede ser menor a inicio"

    if (fin - inicio).days > MAX_DIAS_RANGO:
        return None, f"El rango máximo permitido es {MAX_DIAS_RANGO} días (~6 meses)"

    agrupacion_gastos = args.get("agrupacion_gastos", "proveedor")
    if agrupacion_gastos not in {"proveedor", "categoria"}:
        agrupacion_gastos = "proveedor"

    kpis                       = _calcular_kpis(inicio, fin, id_negocio, col)
    pcts, periodo_anterior_str = _calcular_comparativa(kpis, inicio, fin, id_negocio, col)
    series                     = _cargar_series(inicio, fin, id_negocio, args.get("granularidad", "semana"), col, agrupacion_gastos)

    return {
        **series,
        "periodo_anterior": periodo_anterior_str,
        "kpis":             {**kpis, **pcts},
    }, None
